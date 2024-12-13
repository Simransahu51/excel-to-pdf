from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

def convert_excel_to_pdf(input_excel, output_pdf):
    # Load the Excel workbook
    wb = load_workbook(filename=input_excel)
    ws = wb.active

    # Extract data from the Excel worksheet
    excel_data = []
    for row in ws.iter_rows(values_only=True):
        excel_data.append(row)

    # Create a PDF
    doc = SimpleDocTemplate(output_pdf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Add Excel data to the PDF as a table
    table = Table(excel_data)
    table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('TEXTCOLOR', (0, 0), (-1, 0), (0, 0, 1)),
                                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                                ('INNERGRID', (0, 0), (-1, -1), 0.25, (0, 0, 0)),
                                ('BOX', (0, 0), (-1, -1), 0.25, (0, 0, 0))
                                ]))
    elements.append(table)

    # Build the PDF document
    doc.build(elements)

# Example usage:
if __name__ == "__main__":
    input_excel_file = "C:/Users/simran.kumari/Documents/azurefunction/excelsheet.xlsx"  # Provide the path to your input Excel file
    output_pdf_file = "C:/Users/simran.kumari/Documents/azurefunction/output_folder.pdf"    # Provide the path for the output PDF file

    convert_excel_to_pdf(input_excel_file, output_pdf_file)
